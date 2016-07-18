# -*- coding: utf-8 -*-
from openerp import models, fields, api, exceptions
import base64,xlrd,StringIO
#import xlrd,tkFileDialog,Tkinter
#from io import StringIO
from openpyxl import workbook
from openpyxl.reader.excel  import load_workbook

class PunchTask(models.Model):
    _name = 'punch.task'
    _description = 'Punch task'
    datafile=fields.Binary(u'Excel(.xlsx)',required=True)
    records=fields.Text('有误记录')
    #rawData=xlrd.open_workbook('/home/caofa/test.xls')
    #table=rawData.sheets()[0]
    #x=table.row_values(2)
    @api.multi
    def select_odd(self):
        data_file_p=open('/tmp/test.xlsx','w')                # this method can get data by transportation of a part of disk
        #data_file_p.write((base64.b64decode(self.datafile)))
        data_file_p.write((base64.decodestring(self.datafile)))
        data_file_p.close()
        wb=xlrd.open_workbook(filename=r'/tmp/test.xlsx')
        table=wb.sheets()[0]
        tem=table.row_values(2)
        self.records=tem[2]
'''
        #data=StringIO.StringIO(base64.b64decode(self.datafile)) # thie method can get data by transportation of computer memories and can only read xlsx file rather that xlsx file;
        data=StringIO.StringIO(base64.decodestring(self.datafile))
        wb=load_workbook(data)
        table=wb.worksheets[0]
        tem=table.rows[2]
        self.records=tem[2]
'''
'''
        data_file_p=open('/tmp/test.xlsx','w')                # this method can get data by transportation of a part of disk
        #data_file_p.write((base64.b64decode(self.datafile)))
        data_file_p.write((base64.decodestring(self.datafile)))
        data_file_p.close()
        wb=xlrd.open_workbook(filename=r'/tmp/test.xlsx')
        table=wb.sheets()[0]
        tem=table.row_values(2)
        self.records=tem[2]
'''


    
'''
    def select_odd(self,cr,uid,ids,context=None):
        #try:
        #data=StringIO(self.browse(cr,uid,ids)[0].datafile.decode('utf-8'))
        #wb=load_workbook(data)
        #except:
        #    raise exceptions.Warning(u'错误！'+u'您输入的不是xlsx文件！')
        #ws=wb.worksheets[0]
'''

'''
    #@api.onchange("datafile")
    def select_odd(self,cr,uid,ids,context=None):
        this=self.browse(cr,uid,ids[0])
        #data = base64.decodestring(this.datafile)
        data = this.datafile.decode('base64')
        this.records=data

        with open(data,mode='rb') as f:
            filedata=f.read()
            tem=filedata
            this.records=tem
        #raise exceptions.Warning(self.x)
'''
'''
    @api.multi
    def select_odd(self):
        #root=Tkinter.Tk()
        #dirname=tkFileDialog.askdirectory(parent=root,initialdir="/",title='Please select a directory')
        #if len(dirname)>0:
        #    print "you choose %s" % dirname
        master=Tkinter.Tk()# these two lines delect tk window
        master.withdraw()
      
        filename=tkFileDialog.askopenfilenames(parent=master,filetypes=[("Excel file","*.xls"),("All","*.*")],title="选择文件") # "parent" can confirm to select many times; and if use these codes, when user click the button, poped window which shows to select the target file shold be server rather than user;
        if len(filename)>0:
            rawData=xlrd.open_workbook(filename[0])
            table=rawData.sheets()[0]
            tem=table.row_values(2)
            self.records=tem
'''
    



    




